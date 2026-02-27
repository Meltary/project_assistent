"""Модуль с бизнес-логикой для сравнения файлов 1С и Лоцман."""
import csv
from pathlib import Path
from openpyxl import load_workbook
from typing import Callable, Optional


def read_csv_file_simple(path: str, error_callback: Optional[Callable[[str, str], None]] = None) -> list[list[str]]:
    """
    Чтение CSV с разделителем ';'.
    
    Args:
        path: Путь к CSV файлу
        error_callback: Функция для показа ошибок (title, message)
    
    Returns:
        Список строк файла без полностью пустых строк
        (строки, где все ячейки пустые или из пробелов, отбрасываются).
    """
    rows: list[list[str]] = []
    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f, delimiter=";")
            for r in reader:
                # Пропускаем полностью пустые строки
                if not r or all((c is None or str(c).strip() == "") for c in r):
                    continue
                rows.append(r)
    except Exception as exc:
        if error_callback:
            error_callback("Ошибка чтения CSV", f"{path}\n{exc}")
    return rows


def normalize_string(s: str) -> str:
    """
    Нормализация строки для сравнения.
    
    Args:
        s: Исходная строка
    
    Returns:
        Нормализованная строка (нижний регистр, без лишних пробелов)
    """
    if not s:
        return ""
    return " ".join(s.lower().split())


def compare_files_1c_locman(
    file1_csv: str,
    locman_csvs: list[str],
    error_callback: Optional[Callable[[str, str], None]] = None,
):
    """
    Сравнение CSV 1С и Лоцман (частичное совпадение вторых столбцов).
    
    Args:
        file1_csv: Путь к CSV файлу 1С
        locman_csvs: Список путей к CSV файлам Лоцман
        error_callback: Функция для показа ошибок (title, message)
    
    Returns:
        Кортеж (matches, nomatches, nomatches_1c, locman_options):
        - matches: список (наименование_1с, наименование_лоцман, файл_лоцман)
        - nomatches: список (наименование_лоцман, файл_лоцман)
        - nomatches_1c: список наименований 1С (2-й столбец), которые не нашли совпадений
        - locman_options: словарь (name_loc, loc_path) -> [name_1c, ...]
    """
    data_1c = read_csv_file_simple(file1_csv, error_callback)

    matches: list[tuple[str, str, str]] = []
    nomatches: list[tuple[str, str]] = []
    nomatches_1c: list[str] = []

    # заранее собираем все строки Лоцман (нужны для поиска несовпадений 1С)
    loc_entries: list[tuple[str, str, str]] = []  # (name_loc, key_loc, loc_path)
    for loc_path in locman_csvs:
        data_loc = read_csv_file_simple(loc_path, error_callback)
        for row in data_loc[1:]:
            # Пропускаем строки, содержащие фразу "Куда входит" в любом столбце
            if any("Куда входит" in str(cell) for cell in row if cell is not None):
                continue

            # Должно быть как минимум 5 столбцов, и 5-й столбец (индекс 4) не пустой
            if len(row) < 5:
                continue
            fifth_col = row[4]
            if fifth_col is None or str(fifth_col).strip() == "":
                continue

            if len(row) < 2:
                continue
            name_loc = row[1]  # первый столбец в Лоцмане
            key_loc = normalize_string(name_loc)
            if key_loc:
                loc_entries.append((name_loc, key_loc, loc_path))

    # 1) заполняем совпадения и несовпадения Лоцман
    # также собираем словарь опций для каждого Лоцман
    locman_options: dict[tuple[str, str], list[str]] = {}  # (name_loc, loc_path) -> [name_1c, ...]
    
    for (name_loc, key_loc, loc_path) in loc_entries:
        found = False
        loc_key = (name_loc, loc_path)
        if loc_key not in locman_options:
            locman_options[loc_key] = []
        
        for row1 in data_1c[1:]:
            if len(row1) < 2:
                continue
            name_1c = row1[1]  # первый столбец в 1с
            key_1c = normalize_string(name_1c)
            if not key_1c:
                continue
            if key_loc in key_1c or key_1c in key_loc:
                matches.append((name_1c, name_loc, loc_path))
                if name_1c not in locman_options[loc_key]:
                    locman_options[loc_key].append(name_1c)
                found = True
        if not found:
            nomatches.append((name_loc, loc_path))

    # 2) несовпадения 1С (если строка 1С не нашла совпадений ни с одним Лоцман)
    for row1 in data_1c[1:]:
        if len(row1) < 2:
            continue
        name_1c = row1[1]  # первый столбец в 1с
        key_1c = normalize_string(name_1c)
        if not key_1c:
            continue
        found_1c = False
        for (_name_loc, key_loc, _loc_path) in loc_entries:
            if key_loc in key_1c or key_1c in key_loc:
                found_1c = True
                break
        if not found_1c:
            nomatches_1c.append(name_1c)

    return matches, nomatches, nomatches_1c, locman_options


def convert_xlsx_to_csv(
    path: str,
    error_callback: Optional[Callable[[str, str], None]] = None
) -> str:
    """
    Если файл XLSX, конвертирует в CSV (разделитель ';') и возвращает путь к CSV,
    иначе возвращает исходный путь.
    
    Args:
        path: Путь к файлу
        error_callback: Функция для показа ошибок (title, message)
    
    Returns:
        Путь к CSV файлу
    """
    if not path or not path.lower().endswith(".xlsx"):
        return path

    src = Path(path)
    dst = src.with_suffix(".csv")

    try:
        wb = load_workbook(filename=src, read_only=True, data_only=True)
    except Exception as exc:
        if error_callback:
            error_callback("Ошибка XLSX", str(exc))
        return path

    ws = wb.active  # если нужен другой лист — заменить на wb[sheet_name]

    try:
        with dst.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            for row in ws.iter_rows(values_only=True):
                # Пропускаем полностью пустые строки
                if row is None:
                    continue
                row_list = list(row)
                if all((cell is None or str(cell).strip() == "") for cell in row_list):
                    continue
                # csv.writer корректно обработает None
                writer.writerow(row_list)
    except Exception as exc:
        if error_callback:
            error_callback("Ошибка записи CSV", str(exc))
        return path
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return str(dst)
